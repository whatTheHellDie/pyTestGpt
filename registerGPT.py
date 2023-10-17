import webbrowser    
import pyautogui
import time
from clickImage import clickImage
from clickMove import clickMove
from checkWebsiteLoading import checkWebsiteLoading
import openpyxl
import random
import string
file_name = "data.xlsx"
def mRegisterGPT(mode='user'):
    if mode =='auto':
        clickImage('images/10.png')
        time.sleep(0.5)
    else:
        url = "https://sms-activate.org/cn"
        chrome_path = "C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe"
        webbrowser.register("chrome", None, webbrowser.BackgroundBrowser(chrome_path))
        webbrowser.get("chrome").open_new(url)
        time.sleep(2)
        clickImage('images/0.png')
        clickImage('images/1.png')
        time.sleep(0.5)

    gptUrl = "https://chat.openai.com/auth/login"
    pyautogui.typewrite(gptUrl)
    pyautogui.press('enter') #防止输入法
    pyautogui.press('enter')
    checkWebsiteLoading('images/11.png',next)
    
    
def next():
    time.sleep(2)
    clickImage('images/11.png',0.7)
    pyautogui.click()
    checkWebsiteLoading('images/12.png',loadingWebsiteSuccess)

def loadingWebsiteSuccess():
    
    wb = openpyxl.load_workbook(file_name)
    ws = wb.active  # 或者使用工作表的名字：ws = wb['Sheet1']

    # 找到最后一行
    last_row = ws.max_row

    # 获取最后一行第一列的数据
    email = ws.cell(row=last_row, column=1).value

    

    gptPassword=generate_random_password()
    ws.cell(row=last_row, column=3, value=gptPassword)
    # 一定要记得关闭工作簿
    wb.save(file_name)
    
    pyautogui.typewrite(email)
    clickImage('images/13.png')
    checkWebsiteLoading('images/14.png',accountSuccess)

def accountSuccess():
    wb = openpyxl.load_workbook(file_name)
    ws = wb.active  # 或者使用工作表的名字：ws = wb['Sheet1']
    last_row = ws.max_row
    gptPassword = ws.cell(row=last_row, column=3).value
    pyautogui.typewrite(gptPassword)
    clickImage('images/13.png')


def generate_random_password():
    # 生成前两位的随机大写字母
    upper_letters = ''.join(random.choice(string.ascii_uppercase) for _ in range(2))

    # 生成六位随机小写字母
    lower_letters = ''.join(random.choice(string.ascii_lowercase) for _ in range(8))

    # 生成两位随机数字
    numbers = ''.join(random.choice(string.digits) for _ in range(2))

    # 组合生成的部分以生成最终的字符串
    final_string = upper_letters + lower_letters + numbers

    return final_string