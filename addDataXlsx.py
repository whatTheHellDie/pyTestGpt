import openpyxl
file_name = "data.xlsx"
def initExcel():
    # 文件名
    try:
        # 尝试打开文件，如果文件不存在，会引发异常
        workbook = openpyxl.load_workbook(file_name)
    except FileNotFoundError:
        # 如果文件不存在，创建一个新的工作簿
        workbook = openpyxl.Workbook()

        # 获取工作表
        sheet = workbook.active

        # 数据
        data = [
            ["ProtonMail账号(也是chatgpt账号)", "ProtonMail密码","chatgpt密码"],
        ]

        # 将数据写入工作表
        for row in data:
            sheet.append(row)
        sheet.column_dimensions['A'].width = 50

        # 设置第二列（B列）的宽度为15
        sheet.column_dimensions['B'].width = 50

        # 设置第三列（C列）的宽度为15
        sheet.column_dimensions['C'].width = 50
        # 保存工作簿到文件
        workbook.save(file_name)


    return workbook


# 示例调用addDataXlsx方法，根据 mode 参数执行不同的操作
# data1 = {"newAccount": "User1", "newPassword": "Pass1"}
# update_excel("account", data1)

# data2 = {"newPassword": "Pass2"}
# update_excel("password", data2)
def addDataXlsx(mode, data):
    initExcel()
    # 打开 Excel 文件
    workbook = openpyxl.load_workbook(file_name)

    # 获取工作表
    sheet = workbook.active

    if mode == "account":
        # 如果 mode 为 "account"，增加新行
        new_account = data.get("newAccount", "")
        new_account+='@proton.me'
        new_password = data.get("newPassword", "")
        new_row = [new_account, new_password]
        sheet.append(new_row)
    elif mode == "password":
        # 如果 mode 为 "password"，在最后一行的第三列添加新密码
        new_password = data.get("newPassword", "")
        last_row = sheet.max_row
        sheet.cell(row=last_row, column=3, value=new_password)

    # 保存工作簿
    workbook.save(file_name)

